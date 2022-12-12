# Extracting sum of open hours and total check-out people
import numpy as np
import pandas as pd
import time
import openpyxl
import matplotlib.pyplot as mp

def find_response(current_lanes : np.array, index : int):
    # This function will return after how many 3-min tiem slots will the store staff respond to our open lane alerts
    # if not, return -1
    # TODO: thereshold for response or not
    threshold = 5
    current_lane = current_lanes[index]
    for i in range(1, threshold):
        # print(str(current_lanes[index + i]) + "~~~~" + str(current_lane))
        # time.sleep(1)
        if current_lanes[index + i] > current_lane:
            return i
    return 0
if __name__ == '__main__':

    ###################################################################################################################
    stores_name = ["0801", "0165", "2528", "1960", "2411"]
    basepath = "/home/turingalbertsons1/PycharmProjects/analyzeResponse/"
    # TODO: what dates need to be processed
    dates_name = ["2022-11-28", "2022-11-29", "2022-11-30", "2022-12-01", "2022-12-02", "2022-12-03", "2022-12-04",
                  "2022-12-05", "2022-12-06", "2022-12-07", "2022-12-08", "2022-12-09", "2022-12-10", "2022-12-11"]
    # TODO: heads for first row of data
    heads = ["Site", "Date", "Alert Generated", "Alert Responded", "Response Time", "0~3", "3~6", "6~9", "9~12", "12~15",
             "15~18", "18~21", "21~24", "24~27", "27~30"]
    # TODO: query 15 mins (5 timeslots) before and after alert responded, feel free to tune if needed
    num_test = 5
    # TODO: no multiple alerts within 15 mins, before it's 30 mins
    cool_down = 5
    # TODO: thereshold for open and close lane alert and long queue alert
    alpha_upper = 1.0
    alpha_lower = -2.2
    thereshold_lqa = 3.90

    number_timeslots = 480

    # TODO: number of dimension each lane has, currently it's 4, it may be 5 when queeu time is divided into wait and service
    number_dim = 4
    num_dates = len(dates_name)
    # TODO: csv files in "alert" folder start from 00:30 which is 10 time slots after 00:00
    bias_source_alert = 10
    bias_source_alert_end = 465
    number_lanes = 7
    # TODO: in the out put there are 6 - 1 = 5 columns which describe alert info: site, date, generated, responded, duration
    bias_output_sheet = 6

    wb = openpyxl.Workbook()
    wb.create_sheet("Summary")
    wb.remove_sheet(wb['Sheet'])
    wb.create_sheet("avg_queue_length")
    wb.create_sheet("avg_wait_time")
    wb.create_sheet("avg_number_lqa")
    count = 1
    sheet_queue_length = wb["avg_queue_length"]
    sheet_wait_time = wb["avg_wait_time"]
    sheet_number_lqa = wb["avg_number_lqa"]
    for idx_head in range(len(heads)):
        sheet_queue_length.cell(row=1, column=idx_head + 1).value = heads[idx_head]
        sheet_wait_time.cell(row=1, column=idx_head + 1).value = heads[idx_head]
        sheet_number_lqa.cell(row=1, column=idx_head + 1).value = heads[idx_head]
    for store in stores_name:
        print("~~~ Start processing response at store " + store + " ~~~")
        for date in dates_name:
            temp_d = pd.Timestamp(date)
            if temp_d.dayofweek < 5:
                # TODO: delta for weekday
                delta = 1.75
            else:
                # TODO: delta for weekends
                delta = 1.80
            csv_source = pd.read_csv(basepath + "source/" + store + "/" + date + ".csv")  # usecols=[]
            np_source = np.array(csv_source)
            csv_alerts = pd.read_csv(basepath + "alert/" + store + "/alarm_results-" + store + "-date" + date + ".csv")
            np_alerts = np.array(csv_alerts)

            timeslots = np_source[:, 1]
            curren_lanes = np_source[:, 2]
            status = np_source[:, 4]
            queue_length = np_source[:, 5]
            wait_time = np_source[:, 6]
            customer_throughput = np_source[:, 7]

            predicted_length = np_alerts[:, 4]
            actual_length = np_alerts[:, 5]
            predicted_lanes = np_alerts[:, 6]
            for idx in range(1, number_lanes):
                # print(status.shape)
                # print(queue_length.shape)
                # print(wait_time.shape)
                # print(customer_throughput.shape)
                status = np.vstack((status, np_source[:, 4 + idx * number_dim]))
                queue_length = np.vstack((queue_length, np_source[:, 5 + idx * number_dim]))
                wait_time = np.vstack((wait_time, np_source[:, 6 + idx * number_dim]))
                customer_throughput = np.vstack((customer_throughput, np_source[:, 7 + idx * number_dim]))

            avg_queue_length = np.sum(queue_length, axis=0) / (number_lanes * 1.0)
            avg_wait_time = np.sum(wait_time, axis=0) / (number_lanes * 1.0)
            total_customer_throughput = np.sum(customer_throughput, axis=0)
            long_queue_alerts = np.sum(np.where(queue_length >= thereshold_lqa, 1, 0), axis=0)

            # print(curren_lanes.shape)
            # print(avg_queue_length.shape)
            # print(avg_wait_time.shape)
            # print(total_customer_throughput.shape)
            k = bias_source_alert
            while k < bias_source_alert_end:
                # elif predicted_length[k - bias_source_alert] / delta - curren_lanes[k] <= alpha_lower:
                #     new_sheet.cell(row=count, column=11).value = "-1"
                if predicted_lanes[k - bias_source_alert] - curren_lanes[k] >= alpha_upper:
                    count = count + 1
                    sheet_queue_length.cell(row=count, column=1).value = store
                    sheet_queue_length.cell(row=count, column=2).value = date
                    sheet_queue_length.cell(row=count, column=3).value = timeslots[k][5:]
                    sheet_wait_time.cell(row=count, column=1).value = store
                    sheet_wait_time.cell(row=count, column=2).value = date
                    sheet_wait_time.cell(row=count, column=3).value = timeslots[k][5:]
                    sheet_number_lqa.cell(row=count, column=1).value = store
                    sheet_number_lqa.cell(row=count, column=2).value = date
                    sheet_number_lqa.cell(row=count, column=3).value = timeslots[k][5:]

                    find_idx = find_response(curren_lanes, k)
                    if find_idx: # if there is response
                        sheet_queue_length.cell(row=count, column=4).value = timeslots[k + find_idx][5:]
                        sheet_queue_length.cell(row=count, column=5).value = 3 * find_idx

                        sheet_wait_time.cell(row=count, column=4).value = timeslots[k + find_idx][5:]
                        sheet_wait_time.cell(row=count, column=5).value = 3 * find_idx

                        sheet_number_lqa.cell(row=count, column=4).value = timeslots[k + find_idx][5:]
                        sheet_number_lqa.cell(row=count, column=5).value = 3 * find_idx

                        for l in range(2 * num_test):
                            sheet_queue_length.cell(row=count, column=l + bias_output_sheet).value = avg_queue_length[k + find_idx - num_test + l]
                            sheet_wait_time.cell(row=count, column=l + bias_output_sheet).value = avg_wait_time[k + find_idx - num_test + l]
                            sheet_number_lqa.cell(row=count, column=l + bias_output_sheet).value = long_queue_alerts[k + find_idx - num_test + l]
                    else:
                        for l in range(2 * num_test):
                            sheet_queue_length.cell(row=count, column=l + bias_output_sheet).value = avg_queue_length[k + l]
                            sheet_wait_time.cell(row=count, column=l + bias_output_sheet).value = avg_wait_time[k + l]
                            sheet_number_lqa.cell(row=count, column=l + bias_output_sheet).value = long_queue_alerts[k + l]
                    k = k + cool_down
                else:
                    k = k + 1
    wb.save('./response_analyze.xlsx')
    print("~~~~~~   Finished!   ~~~~~~")