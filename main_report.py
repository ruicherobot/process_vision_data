# Extracting sum of open hours and total check-out people
import numpy as np
import pandas as pd
import time
import openpyxl
import matplotlib.pyplot as mp
import csv
if __name__ == '__main__':
    wb = openpyxl.Workbook()
    new_sheet = wb['Sheet']
    new_sheet.cell(row=1, column=1).value = "Site"
    new_sheet.cell(row=1, column=2).value = "Date"
    new_sheet.cell(row=1, column=3).value = "time_3min"
    new_sheet.cell(row=1, column=4).value = "predicted_next_15min_avg_length"
    new_sheet.cell(row=1, column=5).value = "actual_next_15min_avg_length"
    new_sheet.cell(row=1, column=6).value = "actual_3min_length"
    new_sheet.cell(row=1, column=7).value = "predicted_number_of_lanes_need"
    new_sheet.cell(row=1, column=8).value = "actual_number_of_lanes"
    new_sheet.cell(row=1, column=9).value = "customer_throughput"
    new_sheet.cell(row=1, column=10).value = "open_lane_alerts"
    new_sheet.cell(row=1, column=11).value = "close_lane_alerts"
    alpha_upper = 1.0
    alpha_lower = -2.2
    step_size = 5
    number_timeslots = 480
    number_lanes = 7
    number_dim = 4
    stores_name = ["2411", "2528", "1960", "0165", "0801"]
    count = 1
    export_name = "Data_for_weekly_report.xlsx"
    for store_name in stores_name:
        print("~~~~~~Start processing for Store " + store_name + " ~~~~~~")
        source_name = store_name + "_source.xlsx"
        predicted_name = store_name + "_predicted.xlsx"
        # TODO: Mike wants us to convert xlsx to csv files, this is a switch for this action
        # xlsx_to_csv = True
        source = openpyxl.load_workbook(source_name)
        predicted = openpyxl.load_workbook(predicted_name)
        sheet_all = source.sheetnames
        for datename in sheet_all:
            # delta is different between configuration of weekdays and weekends
            print("  ````Dealing with data on " + datename + " ````")
            temp_d = pd.Timestamp(datename)
            if temp_d.dayofweek < 5:
                # TODO: delta for weekday
                delta = 1.75
            else:
                # TODO: delta for weekends
                delta = 1.80
            filename = datename + ".csv"
            # status = np.zeros((number_timeslots, number_lanes), dtype=float)
            # queue_length = np.zeros((number_timeslots, number_lanes), dtype=float)
            # wait_time = np.zeros((number_timeslots, number_lanes), dtype=float)
            # customer_throughput = np.zeros((number_timeslots, number_lanes), dtype=float)
            rawData_Original = np.array(pd.read_excel(source_name, sheet_name=datename))
            rawData_Predicted = np.array(pd.read_excel(predicted_name, sheet_name=datename))
            
            timeslots = rawData_Original[:, 1]
            curren_lanes = rawData_Original[:, 2]
            status = rawData_Original[:, 4]
            queue_length = rawData_Original[:, 5]
            wait_time = rawData_Original[:, 6]
            customer_throughput = rawData_Original[:, 7]

            predicted_length = rawData_Predicted[:, 1]
            actual_length = rawData_Predicted[:, 2]
            
            for idx in range(1, number_lanes):
                # print(status.shape)
                # print(queue_length.shape)
                # print(wait_time.shape)
                # print(customer_throughput.shape)
                status = np.vstack((status, rawData_Original[:, 4 + idx * number_dim]))
                queue_length = np.vstack((queue_length, rawData_Original[:, 5 + idx * number_dim]))
                wait_time = np.vstack((wait_time, rawData_Original[:, 6 + idx * number_dim]))
                customer_throughput = np.vstack((customer_throughput, rawData_Original[:, 7 + idx * number_dim]))
    
            avg_queue_length = np.sum(queue_length, axis=0) / (number_lanes * 1.0)
            avg_wait_time = np.sum(wait_time, axis=0) / (number_lanes * 1.0)
            total_customer_throughput = np.sum(customer_throughput, axis=0)
            # print(curren_lanes.shape)
            # print(avg_queue_length.shape)
            # print(avg_wait_time.shape)
            # print(total_customer_throughput.shape)
            for k in range(number_timeslots):
                count = count + 1
                new_sheet.cell(row=count, column=1).value = store_name
                new_sheet.cell(row=count, column=2).value = datename
                new_sheet.cell(row=count, column=3).value = timeslots[k]
    
                new_sheet.cell(row=count, column=6).value = avg_queue_length[k]
    
                new_sheet.cell(row=count, column=8).value = curren_lanes[k]
                new_sheet.cell(row=count, column=9).value = total_customer_throughput[k]
    
                # if not np.isnan(predicted_length[k]):
                #     new_sheet.cell(row=count, column=4).value = predicted_length[k]
                #     new_sheet.cell(row=count, column=7).value = predicted_length[k] / delta
                #     new_sheet.cell(row=count, column=5).value = actual_length[k]
                #     if predicted_length[k] / delta - curren_lanes[k] >= alpha_upper:
                #         new_sheet.cell(row=count, column=10).value = "1"
                #     elif predicted_length[k] / delta - curren_lanes[k] <= alpha_lower:
                #         new_sheet.cell(row=count, column=11).value = "-1"

                if 10 <= k < 465:
                    new_sheet.cell(row=count, column=4).value = predicted_length[k - 10]
                    new_sheet.cell(row=count, column=7).value = predicted_length[k - 10] / delta
                    new_sheet.cell(row=count, column=5).value = actual_length[k -10]
                    if predicted_length[k -10] / delta - curren_lanes[k] >= alpha_upper:
                        new_sheet.cell(row=count, column=10).value = "1"
                    elif predicted_length[k - 10] / delta - curren_lanes[k] <= alpha_lower:
                        new_sheet.cell(row=count, column=11).value = "-1"

        # if xlsx_to_csv:
        #     for datename in sheet_all:
        #         filename = "/home/turingalbertsons1/PycharmProjects/analyzeResponse/" + store_name + "/" + datename + ".csv"
        #         print(filename)
        #         sheet = wb[datename]
        #         col = csv.writer(open(filename, 'w', newline=""))
        #         for r in sheet.rows:
        #             col.writerow([cell.value for cell in r])
    wb.save(export_name)
    print(" Finished at " + export_name)